import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import os

file_name= "students.xlsx"

#create Excel file if it doesn't exist
if not os.path.exists(file_name):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["ID","Name","Course","Phone","Grade"])
    wb.save(file_name)

#1 Register student
def register_student():
    student_ID=entry_id.get()
    name=entry_name.get()
    course=entry_course.get()
    phone=entry_phone.get()

    if student_ID=="" or name=="" or course==""or phone=="":
        messagebox.showerror("Error","all fields are required")
        return

    wb= load_workbook(file_name)
    ws=wb.active
    ws.append([student_ID, name, course, phone, ""])
    wb.save(file_name)
    messagebox.showinfo("Success","Student Registerd Successsfully")
    entry_ID.delete(0,tk.END)
    entry_name.delete(0,tk.END)
    entry_course.delete(0,tk.END)
    entry_phone.delete(0,tk.END)

#Display students
def display_student():
    wb = load_workbook(file_name)
    ws = wb.active
    display_window=tk.Toplevel(root)
    display_window.title("Registered Students")

    text=tk.Text(display_window)
    text.pack()

    for row in ws.iter_rows(values_only=True):
        text.insert(tk.END,str(row)+ "\n")

#Assign course+ grade
def assign_grade():
    student_ID=entry_assign_id.get()
    grade= entry_grade.get()

    wb=load_workbook(file_name)
    ws= wb.active

    for row in ws.iter_rows():
        if str(row[0].value) == student_ID:
            row[4].value=grade
            wb.save(file_name)
            messagebox.showinfo("Success","grade assigned successfully")
            return
    messagebox.showerror("Error","student ID not found")

root = tk.Tk()
root.title("School Management System")
root.geometry("400x500")

tk.Label(root, text="Register New Student", font=("Arial", 14)).pack(pady=5)

tk.Label(root, text="ID").pack()
entry_id = tk.Entry(root)
entry_id.pack()

tk.Label(root, text="Name").pack()
entry_name = tk.Entry(root)
entry_name.pack()

tk.Label(root, text="Course").pack()
entry_course = tk.Entry(root)
entry_course.pack()

tk.Label(root, text="Phone").pack()
entry_phone = tk.Entry(root)
entry_phone.pack()

tk.Button(root, text="Register Student", command=register_student).pack(pady=10)

tk.Button(root, text="Display Registered Students", command=display_student).pack(pady=5)

tk.Label(root, text="Assign Grade", font=("Arial", 14)).pack(pady=10)

tk.Label(root, text="Student ID").pack()
entry_assign_id = tk.Entry(root)
entry_assign_id.pack()

tk.Label(root, text="Grade").pack()
entry_grade = tk.Entry(root)
entry_grade.pack()

tk.Button(root, text="assign_grade",command=assign_grade).pack(pady=10)

root.mainloop()